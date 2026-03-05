#!/usr/bin/env python3
"""Generate data/georef_tree_2025.json from official XLSX catalog.

Reads Departamento/Distrito/Ciudad relationships and builds nested maps.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, Tuple

from openpyxl import load_workbook


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text)


def _clean_name(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            value = int(value)
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits


def _set_unique(mapping: Dict[str, str], code: str, name: str, kind: str) -> None:
    if not code or not name:
        return
    prev = mapping.get(code)
    if prev is None:
        mapping[code] = name
        return
    if prev != name:
        raise ValueError(f"Conflicto {kind} {code}: '{prev}' vs '{name}'")


def _find_header(ws) -> Tuple[int, Dict[str, int]]:
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=60, values_only=True), 1):
        normed = [_norm_header(val) for val in row]
        if not normed:
            continue
        if any("departamento" in val for val in normed) and any("distrito" in val for val in normed) and any("ciudad" in val for val in normed):
            dep_name_idx = next(i for i, v in enumerate(normed) if "departamento" in v)
            dist_name_idx = next(i for i, v in enumerate(normed) if "distrito" in v)
            city_name_idx = next(i for i, v in enumerate(normed) if "ciudad" in v)
            if dep_name_idx == 0 or dist_name_idx == 0 or city_name_idx == 0:
                raise ValueError("Encabezado de códigos geográficos inválido (faltan columnas de código).")
            return idx, {
                "dep_code": dep_name_idx - 1,
                "dep_name": dep_name_idx,
                "dist_code": dist_name_idx - 1,
                "dist_name": dist_name_idx,
                "city_code": city_name_idx - 1,
                "city_name": city_name_idx,
            }
    raise RuntimeError("No se encontró encabezado con Departamento/Distrito/Ciudad en el XLSX.")


def build_georef_tree(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    header_row, cols = _find_header(ws)

    dep: Dict[str, str] = {}
    dist_by_dep: Dict[str, Dict[str, str]] = {}
    city_by_dist: Dict[str, Dict[str, str]] = {}
    city_to_dist: Dict[str, str] = {}
    dist_to_dep: Dict[str, str] = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        dep_code_raw = _normalize_code(row[cols["dep_code"]])
        dep_name = _clean_name(row[cols["dep_name"]])
        dist_code_raw = _normalize_code(row[cols["dist_code"]])
        dist_name = _clean_name(row[cols["dist_name"]])
        city_code_raw = _normalize_code(row[cols["city_code"]])
        city_name = _clean_name(row[cols["city_name"]])

        if not (dep_code_raw and dep_name and dist_code_raw and dist_name and city_code_raw and city_name):
            continue

        dep_code = dep_code_raw.zfill(2)
        dist_code = dist_code_raw.zfill(4)
        city_code = city_code_raw.zfill(5)

        _set_unique(dep, dep_code, dep_name, "departamento")

        dist_by_dep.setdefault(dep_code, {})
        _set_unique(dist_by_dep[dep_code], dist_code, dist_name, "distrito")

        city_by_dist.setdefault(dist_code, {})
        _set_unique(city_by_dist[dist_code], city_code, city_name, "ciudad")

        prev_dist = city_to_dist.get(city_code)
        if prev_dist and prev_dist != dist_code:
            raise ValueError(
                f"Conflicto ciudad->distrito {city_code}: '{prev_dist}' vs '{dist_code}'"
            )
        city_to_dist[city_code] = dist_code

        prev_dep = dist_to_dep.get(dist_code)
        if prev_dep and prev_dep != dep_code:
            raise ValueError(
                f"Conflicto distrito->departamento {dist_code}: '{prev_dep}' vs '{dep_code}'"
            )
        dist_to_dep[dist_code] = dep_code

    return {
        "dep": dep,
        "dist_by_dep": dist_by_dep,
        "city_by_dist": city_by_dist,
        "city_to_dist": city_to_dist,
        "dist_to_dep": dist_to_dep,
    }


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Genera georef_tree_2025.json desde XLSX oficial.")
    parser.add_argument(
        "--input",
        default="data/catalogos/codigos-de-referecia-geografica-noviembre-2025.xlsx",
        help="Ruta al XLSX oficial",
    )
    parser.add_argument(
        "--output",
        default="data/georef_tree_2025.json",
        help="Ruta de salida JSON",
    )
    args = parser.parse_args(argv)

    xlsx_path = Path(args.input).resolve()
    if not xlsx_path.exists():
        print(f"XLSX no encontrado: {xlsx_path}", file=sys.stderr)
        return 2

    tree = build_georef_tree(xlsx_path)
    out_path = Path(args.output).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(tree, ensure_ascii=False, indent=2), encoding="utf-8")

    dep_count = len(tree["dep"])
    dist_count = sum(len(v) for v in tree["dist_by_dep"].values())
    city_count = sum(len(v) for v in tree["city_by_dist"].values())
    print(f"OK: {dep_count} departamentos, {dist_count} distritos, {city_count} ciudades -> {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
